#!/usr/bin/env python
# -*- coding:utf-8 -*-

from TwitterAPI import TwitterAPI
from TwitterAPI import TwitterError
from requests_oauthlib import OAuth1Session
import json
import tweepy
import oauth2
import re
import datetime
from time import sleep
import time
import pandas as pd
import traceback
import calendar
import openpyxl
import schedule
import csv

excel_column = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
months = {'': 0, 'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}

vote_column = ["統計日時", "検索数", "最小ツイートID", "最小日時", "最大ツイートID", "最大日時", "ヨコハマ", "シブヤ", "ナゴヤ", "シンジュク", "イケブクロ", "オオサカ"]
vote_df = pd.DataFrame(columns=vote_column)





oath_key_dict = {
    "consumer_key": "XX",
    "consumer_secret": "XX",
    "access_token": "XX",
    "access_token_secret": "XX"
}

def fcn_tweet(text):
    url = "https://api.twitter.com/1.1/statuses/update.json?status={}".format(text)

    #consumer = oauth2.Consumer(key='手順３で取得できるConsumer API keys の API key', secret='手順３で取得できるConsumer API keys の API secret key')
    #token = oauth2.Token(key='手順３で取得できるAccess token', secret='手順３で取得できるAccess token secret')
    consumer = oauth2.Consumer(key=CONSUMER_KEY, secret=CONSUMER_SECRET_KEY)
    token = oauth2.Token(key=ACCESS_TOKEN, secret=ACCESS_TOKEN_SECRET)
    client = oauth2.Client(consumer, token)
    resp, content = client.request( url, method="POST")

    return resp, content


def createTweeter():
    # Twitterオブジェクトの生成
    auth = tweepy.OAuthHandler(CONSUMER_KEY, CONSUMER_SECRET_KEY)
    auth.set_access_token(ACCESS_TOKEN, ACCESS_TOKEN_SECRET)
    api = tweepy.API(auth)
    # -------------------------------------------------------------------------
    # 単純にテキスト投稿
    #api.update_status("tweetを投稿 ")
    # 画像付きのツイートを投稿
    #api.update_with_media(status='This is 画像付きツイートテスト from python', filename='C:\\Users\\103408\\Desktop\\pythonAPITest.JPG')

    # 複数画像投稿
    file_names = ["C:\\Users\\103408\\Desktop\\drb_fc_long.jpg", "C:\\Users\\103408\\Desktop\\drb_fc_round.jpg"]
    media_ids = []
    for filename in file_names:
        res = api.media_upload(filename)
        media_ids.append(res.media_id)
    # tweet with multiple images
    # テキスト
    textLine = "ヒプノシスマイク　ヒプマイ HYPSTER ロング缶バッジ トレ缶 丸缶　交換\n譲:寂雷(ロング)、簓(丸)、零、二郎(丸)、三郎\n求:左馬刻、銃兎\n現状でロング缶、丸缶各40個予約済(追加可能性あり)\n郵送お取引きよろしくお願いいたします"
    #api.update_status(status=textLine, media_ids=media_ids)
    '''
    Account = "眠気"  # 取得したいユーザーのユーザーIDを代入
    tweets = api.user_timeline(Account, count=200, page=1)
    num = 1  # ツイート数を計算するための変数
    for tweet in tweets:
        print('twid : ', tweet.id)  # tweetのID
        print('user : ', tweet.user.screen_name)  # ユーザー名
        print('date : ', tweet.created_at)  # 呟いた日時
        print(tweet.text)  # ツイート内容
        print('favo : ', tweet.favorite_count)  # ツイートのいいね数
        print('retw : ', tweet.retweet_count)  # ツイートのリツイート数
        print('ツイート数 : ', num)  # ツイート数
        print('=' * 80)  # =を80個表示
        num += 1  # ツイート数を計算
    '''
    print("===================================================-")

def getDirectMessage():
    token = create_oath_session(oath_key_dict)
    params = {"count": "100"}
    getlist = token.get("https://api.twitter.com/1.1/direct_messages/events/list.json?", params=params)

    dmlist = json.loads(getlist.text)
    #print(dmlist["next_cursor"])

    for line in dmlist["events"]:
        #print(line["text"])
        print(line)

def func_WriteToExcel(se):
    excelName = "DRB_Twitter_Vote.xlsx"
    wb = openpyxl.load_workbook(excelName)
    #print("シート名は" + wb.sheetnames)
    ws = wb["Sheet1"]
    print("最大行数は" + str(ws.max_row))
    print("最大列数は" + str(ws.max_column))

    #wb.save(excelName)

    # ws = wb.create_sheet(title="LicenseInfo")
    try:
        if not se.empty:
            se_value = se.values
            rowNum = ws.max_row + 1
            for index in range(0, len(se_value)):
                cellPos = excel_column[index] + str(rowNum)
                cell = ws[cellPos]
                if cell.value is None:
                    cell.value = se_value[index]

            wb.save(excelName)
    except Exception as e:
        print("統計結果書き込む失敗")
        print(str(Exception))
        print(str(e))
        print(traceback.print_exc())

def getMaxId():
    try:
        with open('./max_id_conf.json', 'r', encoding='utf-8-sig') as load_f:
            load_dict = json.load(load_f)
            currentSearchTwitter = load_dict["CurrentSearchTwitter"]
            max_id = currentSearchTwitter["MaxId"]
            max_id_time = currentSearchTwitter["MaxIDTime"]
            if max_id and max_id_time:
                # すでに検索済みのツイートのidで最も小さいID、当該ツイートの時間
                return max_id, max_id_time
            else:
                return None, None
    except Exception as e:
        print("max_id取得失敗")
        print(str(Exception))
        print(str(e))
        print(traceback.print_exc())

# ツイート時間のフォーマット変換
def datetimeFormatTrans(orig_twitter_datetime):
    rt_val = None
    try:
        if orig_twitter_datetime:
            res = re.split(" ", orig_twitter_datetime)
            if res and len(res) >= 6:
                year_val = res[5]
                month_val = months[res[1]]
                date_val = res[2]
                time_val = res[3]

            rt_val = year_val + "-" + str(month_val) + "-" + date_val + " " + time_val
            return rt_val
    except Exception as e:
        print("ツイート時間フォーマット変換失敗")
        print(str(Exception))
        print(str(e))
        print(traceback.print_exc())

def updateMaxId(max_id, max_id_time):
    try:
        if max_id and max_id_time:
            data_dict = {'CurrentSearchTwitter': {'MaxId': max_id, 'MaxIDTime': max_id_time}}
            # write back to json file
            with open('./max_id_conf.json', 'w', encoding='utf-8-sig') as dump_f:
                json.dump(data_dict, dump_f)
                print("max_id更新成功")
        else:
            print("更新なし")
    except Exception as e:
        print("max_id更新失敗")
        print(str(Exception))
        print(str(e))
        print(traceback.print_exc())

### Functions
def main():

    #resp, content = fcn_tweet("from python again")
    #print("resp : ", resp)
    #print("content : ", content)

    # 新規投稿
    #createTweeter()

    # DM取得
    #getDirectMessage()

    want_pattern = re.compile(r".+(求.+左馬刻)|(求\)左馬刻)|(求】左馬刻)|(求：左馬刻)|(求:左馬刻)|(求\S+左馬刻).*", re.DOTALL)
    result_num = 0
    # MTC
    want_chara_num = 0
    # FP
    want_chara_num1 = 0
    # BAT
    want_chara_num2 = 0
    # MTR
    want_chara_num3 = 0
    # BB
    want_chara_num4 = 0
    # どついたれ本舗
    want_chara_num5 = 0

    search_times = 1
    # すでに検索済みのツイートのidで最も小さいID
    [max_id, max_id_time] = getMaxId()
    max_id = None
    # 今回検索の最新ツイートのidと時間
    current_id = None
    current_id_time = None

    # search_times <=2
    while(True ):
        [tweets, error_code] = tweet_search("ヒプノシスマイク 　2nd D.R.B 投票しました", oath_key_dict, max_id)
        # 改行マッチ（https://www.cnblogs.com/dong973711/p/11924090.html）

        if error_code == 200 and tweets:
            for tweet in tweets["statuses"]:
                print("-----------------------------------------")
                tweet_id = tweet[u'id_str']
                text = tweet['text']

                check_RT = re.findall("RT", text)
                if check_RT:
                    #リツイートは無視
                    continue

                if not current_id:
                    current_id = int(tweet["id"])

                max_id = int(tweet["id"]) - 1

                temp_res = re.findall('“MAD TRIGGER CREW”に投票しました！', text)
                #temp_res = want_pattern.search(text)
                if temp_res:
                    #print("MAD TRIGGER CREWに投票しました")
                    want_chara_num = want_chara_num + 1

                temp_res = re.findall("Fling Posse", text)
                if temp_res:
                    #print("Fling Posseに投票しました")
                    want_chara_num1 += 1

                temp_res = re.findall("Bad Ass Temple", text)
                if temp_res:
                    #print("Bad Ass Templeに投票しました")
                    want_chara_num2 += 1

                temp_res = re.findall("麻天狼", text)
                if temp_res:
                    #print("麻天狼に投票しました")
                    want_chara_num3 += 1

                temp_res = re.findall("Buster Bros!!!", text)
                if temp_res:
                    #print("Buster Bros!!!に投票しました")
                    want_chara_num4 += 1

                temp_res = re.findall("どついたれ本舗", text)
                if temp_res:
                    #print("どついたれ本舗に投票しました")
                    want_chara_num5 += 1

                created_at = tweet[u'created_at']
                user_id = tweet[u'user'][u'id_str']
                user_description = tweet[u'user'][u'description']
                screen_name = tweet[u'user'][u'screen_name']
                user_name = tweet[u'user'][u'name']

                # max_id_time更新
                max_id_time = datetimeFormatTrans(created_at)
                # 今回検索最新のツイートの時間を更新
                if not current_id_time:
                    current_id_time = datetimeFormatTrans(created_at)


                result_num=result_num+1
                print("tweet_id:", tweet_id)
                print("text:", text)
                print("created_at:", created_at)
                #print("user_id:", user_id)
                #print("user_desc:", user_description)
                #print("screen_name:", screen_name)
                print("user_name:", user_name)

                if 'media' in tweet['entities']:
                    medias = tweet['entities']['media']
                    for media in medias:
                        print("midia = ", media['url'])
                        break
                elif 'urls' in tweet['entities']:
                    urls = tweet['entities']['urls']
                    for url in urls:
                        print("url = ", url['url'])
                        break
                else:
                    continue

            search_times += 1
        elif error_code in (420, 429):
            break

            # 結果出力
            print("all result fot search is , ", result_num)
            print("MTCに投票数：", want_chara_num)
            print("FPに投票数：", want_chara_num1)
            print("BATに投票数：", want_chara_num2)
            print("麻天狼に投票数：", want_chara_num3)
            print("BBに投票数：", want_chara_num4)
            print("どついたれ本舗に投票数：", want_chara_num5)

            # 更新max_id
            updateMaxId(max_id, max_id_time)
            print("max_id = %s, max_id = %s" % (str(max_id), max_id_time))
            print("current_id = %s, current_id_time = %s" % (current_id, current_id_time))

            # 結果をexcelファイルに出力
            now = datetime.datetime.now()
            now_str = now.strftime("%Y/%m/%d %H:%M:%S")
            se = pd.Series([now_str, result_num, str(max_id), max_id_time, current_id, current_id_time, want_chara_num,
                            want_chara_num1, want_chara_num2, want_chara_num3, want_chara_num4, want_chara_num5],
                           vote_column)
            event_log_df = vote_df.append(se, vote_column)

            func_WriteToExcel(se)

            # 待ち
            now = datetime.datetime.now()
            print(now.strftime("%Y/%m/%d %H:%M:%S") + ' 接続上限のため待機')
            resetTime = limit_check(oath_key_dict)
            resetTime_datetime = datetime.datetime.strptime(resetTime, '%Y-%m-%d %H:%M:%S')
            waitTime = resetTime_datetime - now
            waitTime_second = waitTime.seconds + 5
            print("%sまで待つ、%d秒"%(resetTime, waitTime_second))

            #sleep(15 * 60)  # 15分待機
            sleep(waitTime_second)
            #continue
        else:
            '''
            print("all result fot search is , ", result_num)
            print("MTCに投票数：", want_chara_num)
            print("FPに投票数：", want_chara_num1)
            print("BATに投票数：", want_chara_num2)
            print("麻天狼に投票数：", want_chara_num3)
            print("BBに投票数：", want_chara_num4)
            print("どついたれ本舗に投票数：", want_chara_num5)
            '''
            break

    # 制限状況を取得
    limit_check(oath_key_dict)

    # 結果出力
    print("all result fot search is , ",result_num)
    print("MTCに投票数：", want_chara_num)
    print("FPに投票数：", want_chara_num1)
    print("BATに投票数：", want_chara_num2)
    print("麻天狼に投票数：", want_chara_num3)
    print("BBに投票数：", want_chara_num4)
    print("どついたれ本舗に投票数：", want_chara_num5)

    # 更新max_id
    updateMaxId(max_id, max_id_time)
    print("max_id = %s, max_id_time = %s" % (str(max_id), max_id_time))
    print("current_id = %s, current_id_time = %s" % (current_id, current_id_time))

    # 結果をexcelファイルに出力
    now = datetime.datetime.now()
    now_str = now.strftime("%Y/%m/%d %H:%M:%S")
    se = pd.Series([now_str, result_num, str(max_id), max_id_time, current_id, current_id_time, want_chara_num, want_chara_num1, want_chara_num2, want_chara_num3, want_chara_num4, want_chara_num5], vote_column)
    event_log_df = vote_df.append(se, vote_column)

    func_WriteToExcel(se)

    return


def create_oath_session(oath_key_dict):
    oath = OAuth1Session(
    oath_key_dict["consumer_key"],
    oath_key_dict["consumer_secret"],
    oath_key_dict["access_token"],
    oath_key_dict["access_token_secret"]
    )

    return oath

# API制限を確認
def limit_check(oath_key_dict):
    url = "https://api.twitter.com/1.1/application/rate_limit_status.json"
    params = {
        "resources": "search"
    }
    # 認証処理
    twitter_instance = create_oath_session(oath_key_dict)
    # 制限チェック処理
    responce = twitter_instance.get(url, params=params)
    response_text = json.loads(responce.text)
    limit_value = response_text["resources"]["search"]["/search/tweets"]["limit"]
    remaining_value = response_text["resources"]["search"]["/search/tweets"]["remaining"]
    reset_time_epoch = response_text["resources"]["search"]["/search/tweets"]["reset"]
    # reset_time_epochをロカール時間に変更
    time_local = time.localtime(reset_time_epoch)
    dateTimeTokyo = time.strftime("%Y-%m-%d %H:%M:%S", time_local)
    print("limit = %d, remaining = %d, reset_time_epoch = %d, reset_time = %s" %(limit_value, remaining_value, reset_time_epoch, dateTimeTokyo))
    return dateTimeTokyo

def tweet_search(search_word, oath_key_dict, max_id):
    try:
        error_code = 0;

        url = "https://api.twitter.com/1.1/search/tweets.json?"
        #url = "https://api.twitter.com/1.1/tweets/search/fullarchive/searchTweet.json"
        params = {
            "q": search_word,
            #"lang": "ja",
            "result_type": "recent",
            "count": "100",
            "max_id": max_id
            }

        # 認証処理
        twitter_instance = create_oath_session(oath_key_dict)
        # 検索操作
        responce = twitter_instance.get(url, params = params)
        error_code = responce.status_code

        # 捜査結果チェック
        if responce.status_code == 200:
            #　無事に検索完了
            tweets = json.loads(responce.text)
            return tweets, error_code

        # APIの接続上限を超えた場合の処理
        # 15分間で180回しかリクエストを送信出来ないので上限に達していたら待つ
        if responce.status_code in (420, 429):
            print("Error code: %d" %(error_code))
            return None, error_code

    except Exception as e:
            print("tweet_search 処理関数実行失敗, error_code = %d"%(error_code))
            print(str(Exception))
            print(str(e))
            print(traceback.print_exc())




# fullarchiveを利用

def searchWordsRecurrent(from_date = None, to_date = None, res = None):
    url = "https://api.twitter.com/1.1/tweets/search/fullarchive/searchVoteTweet.json"
    keyword = "ヒプノシスマイク –Division Rap Battle- 2nd D.R.B"
    print('----------------------------------------------------')
    params = {'query' : keyword, 'maxResults' : 100,'fromDate':from_date,'toDate':to_date}

    twitter_instance = create_oath_session(oath_key_dict)

    #レスポンスが引数で与えられていたら
    if res is not None:
        params['next'] = res['next']

    headers = {'content-type': 'application/json'}
    result = twitter_instance.get(url, headers=headers, params = params)

    #CSVのヘッダーを定義
    header = ['id','User Name','User ID','Follows','Followers','User Location','content','time']
    search_timeline = {}
    if result.status_code == 200:
         with open('data/{keyword}from{from_date}_to{to_date}.csv'.format(keyword = keyword, from_date = from_date, to_date = to_date), 'w') as f:
             search_timeline = json.loads(result.text)
             writer = csv.writer(f)
             writer.writerow(header)
             for tweet in search_timeline['results']:
                 tmp = []
                 tmp.append(tweet['id'])
                 tmp.append(tweet['user']['name'])
                 tmp.append(tweet['user']['screen_name'])
                 tmp.append(tweet['user']['friends_count'])
                 tmp.append(tweet['user']['followers_count'])
                 tmp.append(tweet['user']['location'])
                 tmp.append(tweet['text'])
                 tmp.append(tweet['created_at'])
                 writer.writerow(tmp)
                 tmp = []
             print(len(search_timeline['results']))
    else:
        print("ERROR: %d" % result.status_code)
    if 'next' in search_timeline:
        searchWordsRecurrent(from_date, to_date,search_timeline)

    return

### Execute

schedule.every(20).minutes.do(main)
if __name__ == "__main__":
    #main()
    limit_check(oath_key_dict)
    main()
    #searchWordsRecurrent()

    while True:
        schedule.run_pending()
        time.sleep(1)
